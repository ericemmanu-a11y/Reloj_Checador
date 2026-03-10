import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, time
import os

def init_styles():
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    title_font = Font(size=18, bold=True, color="2c3e50")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Colores corporativos Elite
    header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
    red_font = Font(color="c0392b", bold=True)
    green_font = Font(color="27ae60", bold=True)
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return border, title_font, header_font, header_fill, red_font, green_font, center_align

def format_sheet_header(ws, title, start_date, end_date, num_cols, title_font):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell = ws.cell(row=1, column=1)
    cell.value = title
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:P2')
    ws['A2'].value = f"Del {start_date.strftime('%Y-%m-%d')} al {end_date.strftime('%Y-%m-%d')}"
    ws['A2'].font = Font(bold=True, color="2980b9", size=14)
    ws.row_dimensions[2].height = 25

def check_lateness(t_in, expected_in):
    if t_in > expected_in:
        return (datetime.combine(datetime.min, t_in) - datetime.combine(datetime.min, expected_in)).seconds // 60
    return 0

def check_early_leave(t_out, expected_out):
    if t_out < expected_out:
        return (datetime.combine(datetime.min, expected_out) - datetime.combine(datetime.min, t_out)).seconds // 60
    return 0

def get_spanish_weekday(date_obj):
    dias = {0: "LUN", 1: "MAR", 2: "MIE", 3: "JUE", 4: "VIE", 5: "SAB", 6: "DOM"}
    return dias[date_obj.weekday()]

def create_turnos_sheet(wb, df, border, header_font, header_fill, center_align, s_date, e_date, turnos_manager):
    ws = wb.create_sheet("Reporte de Turnos")
    dates = pd.date_range(s_date, e_date).date
    format_sheet_header(ws, "MATRIZ DE ASISTENCIA DIARIA", s_date, e_date, 3 + len(dates), Font(size=18, bold=True))
    
    headers = ['ID', 'Nombre', 'Turno Base']
    for d in dates:
        headers.append(f"{d.day}\n{get_spanish_weekday(d)}")
    
    ws.append(headers)
    ws.row_dimensions[3].height = 35
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = center_align

    row_idx = 4
    for _, user in df.groupby('ID'):
        uid = str(user['ID'].iloc[0])
        name = user['Nombre'].iloc[0]
        
        # Consultar Turno para embellecer
        tid = turnos_manager.obtener_asignacion(uid)
        t_nombre = turnos_manager.obtener_turnos().get(tid, {}).get('nombre', 'Estándar 09-18') if tid else 'Estándar 09-18'
        
        row_data = [uid, name, t_nombre]
        user_dates = user['Fecha'].unique()
        
        for d in dates:
            row_data.append("Asiste" if d in user_dates else "")
            
        ws.append(row_data)
        for col in range(1, len(row_data) + 1):
            c = ws.cell(row=row_idx, column=col)
            c.border = border
            c.alignment = center_align
        ws.row_dimensions[row_idx].height = 25
        row_idx += 1

    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    for col_idx in range(4, len(headers)+1):
        ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = 10
    ws.freeze_panes = "D4"

def create_estadistico_sheet(wb, df, border, header_font, header_fill, center_align, s_date, e_date, turnos_manager):
    ws = wb.create_sheet("Reporte Estadístico")
    format_sheet_header(ws, "ESTADÍSTICAS GLOBALES DEL PERIODO", s_date, e_date, 15, Font(size=18, bold=True))
    
    headers = [
        'ID', 'Nombre Completo', 'Turno Base', 'Log de Horas Trab.',
        'Total Retardos', 'Mins Tarde', 'Total Salidas Falt/Temp',
        'Mins Perdidos', 'Días Checados', 'Faltas / Ausencias', 'T. Perdido General'
    ]
    
    ws.append(headers)
    ws.row_dimensions[3].height = 40
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = center_align

    total_days = (e_date - s_date).days + 1
    
    row_idx = 4
    for _, user in df.groupby('ID'):
        uid = str(user['ID'].iloc[0])
        name = user['Nombre'].iloc[0]
        
        expected_in, expected_out = turnos_manager.get_turno_por_usuario(uid)
        if not expected_in: expected_in = time(9, 0)
        if not expected_out: expected_out = time(18, 0)
        
        tid = turnos_manager.obtener_asignacion(uid)
        t_nombre = turnos_manager.obtener_turnos().get(tid, {}).get('nombre', 'Estándar 09-18') if tid else 'Estándar 09-18'

        total_hours = 0
        late_count = 0
        late_mins = 0
        early_count = 0
        early_mins = 0
        asistencias = len(user['Fecha'].unique())
        faltas = total_days - asistencias
        
        daily_group = user.groupby('Fecha')
        for current_date, records in daily_group:
            t_in = records['Fecha_Hora'].min().time()
            t_out = records['Fecha_Hora'].max().time()
            
            if len(records) > 1:
                diff = datetime.combine(current_date, t_out) - datetime.combine(current_date, t_in)
                total_hours += diff.total_seconds() / 3600
                
            late = check_lateness(t_in, expected_in)
            if late > 0:
                late_count += 1
                late_mins += late
                
            early = check_early_leave(t_out, expected_out)
            if early > 0 and len(records) > 1:
                early_count += 1
                early_mins += early

        h_int = int(total_hours)
        m_int = int((total_hours - h_int) * 60)
        
        row_data = [
            uid, name, t_nombre, f"{h_int:02d}H {m_int:02d}M",
            late_count, late_mins, early_count, early_mins,
            asistencias, faltas, (late_mins + early_mins)
        ]
        
        ws.append(row_data)
        for col in range(1, len(row_data) + 1):
            c = ws.cell(row=row_idx, column=col)
            c.border = border
            c.alignment = center_align
            # Destaques
            if col in (6, 8, 10) and row_data[col-1] > 0:
                 c.font = Font(color="c0392b", bold=True)
                 
        ws.row_dimensions[row_idx].height = 30
        row_idx += 1

    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    for c in ['E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[c].width = 16

def create_asistencia_sheet(wb, df, border, header_font, header_fill, center_align, s_date, e_date):
    ws = wb.create_sheet("Reporte de Asistencia")
    dates = pd.date_range(s_date, e_date).date
    format_sheet_header(ws, "DETALLE DE FLUJO DE ASISTENCIA", s_date, e_date, 2 + len(dates), Font(size=18, bold=True))
    
    headers = ['ID', 'Empleado']
    for d in dates:
        headers.append(f"{d.day}")
        
    ws.append(headers)
    ws.row_dimensions[3].height = 30
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = center_align

    row_idx = 4
    for _, user in df.groupby('ID'):
        uid = str(user['ID'].iloc[0])
        name = user['Nombre'].iloc[0]
        
        row_data = [uid, name]
        
        for d in dates:
            recs = user[user['Fecha'] == d]
            if len(recs) == 0:
                row_data.append("")
            elif len(recs) == 1:
                row_data.append(recs['Fecha_Hora'].iloc[0].strftime('%H:%M'))
            else:
                t_in = recs['Fecha_Hora'].min().strftime('%H:%M')
                t_out = recs['Fecha_Hora'].max().strftime('%H:%M')
                row_data.append(f"{t_in}\n{t_out}")
                
        ws.append(row_data)
        for col in range(1, len(row_data) + 1):
            c = ws.cell(row=row_idx, column=col)
            c.border = border
            c.alignment = center_align
            # Highlight absent/single scans if needed
        
        ws.row_dimensions[row_idx].height = 40
        row_idx += 1

    ws.column_dimensions['B'].width = 30
    for col_idx in range(3, len(headers)+1):
        ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = 10
    ws.freeze_panes = "C4"

def create_excepciones_sheet(wb, df, border, header_font, header_fill, center_align, red_font, s_date, e_date, turnos_manager):
    ws = wb.create_sheet("Reporte de Excepciones")
    format_sheet_header(ws, "LISTADO CRÍTICO DE EXCEPCIONES Y RETARDOS", s_date, e_date, 10, Font(size=18, bold=True))
    
    headers = ['ID', 'Trabajador', 'Turno Aplicado', 'Día Registrado', 'Check IN', 'Check OUT', 'Retardo (Min)', 'Fuga T. (Min)', 'Omisión Salida', 'Análisis Algorítmico']
    ws.append(headers)
    ws.row_dimensions[3].height = 40
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = center_align
        
    row_idx = 4
    for _, user in df.groupby(['ID', 'Fecha']):
        uid = str(user['ID'].iloc[0])
        name = user['Nombre'].iloc[0]
        date_val = user['Fecha'].iloc[0]
        
        expected_in, expected_out = turnos_manager.get_turno_por_usuario(uid)
        if not expected_in: expected_in = time(9, 0)
        if not expected_out: expected_out = time(18, 0)
        
        tid = turnos_manager.obtener_asignacion(uid)
        t_nombre = turnos_manager.obtener_turnos().get(tid, {}).get('nombre', 'Estándar') if tid else 'Estándar'
        
        t_in = user['Fecha_Hora'].min().time()
        t_out = user['Fecha_Hora'].max().time()
        
        t_in_str = t_in.strftime('%H:%M')
        t_out_str = t_out.strftime('%H:%M') if len(user) > 1 else "NO M."
        
        late = check_lateness(t_in, expected_in)
        early = check_early_leave(t_out, expected_out) if len(user) > 1 else 0
        faltan_datos = 1 if len(user) == 1 else 0
        
        if late == 0 and early == 0 and faltan_datos == 0:
            continue # Clean sheet - only exceptions!

        notas = ""
        if faltan_datos: notas += "[ALERTA] Olvido marcar salida. "
        if late > 0: notas += f"Penalización Tarde: +{late} min. "
        if early > 0: notas += f"Fuga Laboral: +{early} min. "

        row_data = [uid, name, t_nombre, date_val.strftime('%d/%m/%Y'), t_in_str, t_out_str,
                    late, early, "Comprobada" if faltan_datos else "No", notas]
                    
        ws.append(row_data)
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row_idx, column=col)
            c.border = border
            c.alignment = center_align
            if col == 10 and faltan_datos:
                c.font = Font(color="c0392b", bold=True)
        ws.row_dimensions[row_idx].height = 25
        row_idx += 1
        
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['J'].width = 60

def export_to_excel(registros, nombre_archivo, valid_user_ids, turnos_manager):
    if not registros:
        return False, "No hay registros en la memoria del reloj."
        
    df = pd.DataFrame(registros)
    df = df[df['ID'].isin(valid_user_ids)]
    
    if df.empty:
        return False, "Las asistencias presentes no pertenecen a la plantilla cargada vigente."
        
    df['Fecha'] = df['Fecha_Hora'].dt.date
    df['Hora'] = df['Fecha_Hora'].dt.time
    
    s_date = df['Fecha'].min()
    e_date = df['Fecha'].max()
    
    wb = Workbook()
    
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    border, title_font, header_font, header_fill, red_font, green_font, center_align = init_styles()

    create_turnos_sheet(wb, df, border, header_font, header_fill, center_align, s_date, e_date, turnos_manager)
    create_estadistico_sheet(wb, df, border, header_font, header_fill, center_align, s_date, e_date, turnos_manager)
    create_asistencia_sheet(wb, df, border, header_font, header_fill, center_align, s_date, e_date)
    create_excepciones_sheet(wb, df, border, header_font, header_fill, center_align, red_font, s_date, e_date, turnos_manager)

    reportes_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    
    full_path = os.path.join(reportes_dir, nombre_archivo)
    
    try:
        wb.save(full_path)
        return True, full_path
    except Exception as e:
        return False, str(e)
